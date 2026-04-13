from __future__ import annotations

import calendar
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from bson import ObjectId
from openpyxl.styles import Alignment, Font


EXPORT_DIR = Path(__file__).resolve().parent / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _safe_class_name(class_name: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "-" for ch in class_name.strip())


def _as_str_id(value: Any) -> str:
    return "" if value is None else str(value)


def _coerce_object_id(value: Any):
    try:
        text = str(value)
    except Exception:
        return None
    return ObjectId(text) if ObjectId.is_valid(text) else None


def _extract_day(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str) and value:
        return value[:10]
    return ""


def _extract_time(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, str) and value:
        return value[11:19] if len(value) >= 19 and "T" in value else value
    return "-"


def _normalize_status(record: Dict[str, Any]) -> str:
    status = str(record.get("status") or "present").strip().lower()
    if status in ("present", "late"):
        return "Present"
    if status == "absent":
        return "Absent"
    return "Present"


def _apply_sheet_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def _resolve_class_ids(db, class_selector: str) -> List[str]:
    classes = db.get_collection("classes")
    selector = (class_selector or "").strip()
    if not selector:
        return []

    or_filters: List[Dict[str, Any]] = [
        {"name": selector},
        {"classId": selector},
        {"grade": selector},
        {"name": {"$regex": f"^{selector}$", "$options": "i"}},
        {"classId": {"$regex": f"^{selector}$", "$options": "i"}},
    ]

    as_oid = _coerce_object_id(selector)
    if as_oid is not None:
        or_filters.append({"_id": as_oid})

    docs = list(classes.find({"$or": or_filters}, {"_id": 1}))
    return [str(doc.get("_id")) for doc in docs if doc.get("_id")]


def _fetch_students_for_class(db, class_selector: str, class_ids: List[str]):
    students_coll = db.get_collection("students")
    history_coll = db.get_collection("studentclasshistories")

    roster: List[Dict[str, Any]] = []
    if class_ids:
        class_oids = [oid for oid in (_coerce_object_id(cid) for cid in class_ids) if oid is not None]
        hist_filter: Dict[str, Any] = {"status": "active"}
        if class_oids:
            hist_filter["classId"] = {"$in": class_oids}
        else:
            hist_filter["classId"] = {"$in": class_ids}

        histories = list(history_coll.find(hist_filter, {"studentId": 1, "rollNo": 1}).sort("updatedAt", -1))
        student_ids = []
        roll_by_student: Dict[str, str] = {}
        for history in histories:
            sid = _as_str_id(history.get("studentId"))
            if not sid:
                continue
            if sid not in roll_by_student:
                roll_by_student[sid] = str(history.get("rollNo") or "")
                student_ids.append(sid)

        if student_ids:
            student_oids = [oid for oid in (_coerce_object_id(sid) for sid in student_ids) if oid is not None]
            q: Dict[str, Any] = {"_id": {"$in": student_oids}} if student_oids else {"_id": {"$in": student_ids}}
            for student in students_coll.find(q, {"name": 1}):
                sid = _as_str_id(student.get("_id"))
                roster.append(
                    {
                        "_id": sid,
                        "name": student.get("name") or sid,
                        "rollNo": roll_by_student.get(sid) or "-",
                    }
                )

    if roster:
        return roster

    # Legacy fallback for older schemas.
    q = {
        "$or": [
            {"class": class_selector},
            {"class_name": class_selector},
            {"grade": class_selector},
            {"classId": class_selector},
        ]
    }
    return list(students_coll.find(q, {"name": 1, "rollNo": 1}))


def _build_class_filter(class_ids: List[str]) -> Dict[str, Any]:
    if not class_ids:
        return {}

    class_oids = [oid for oid in (_coerce_object_id(cid) for cid in class_ids) if oid is not None]
    variants: List[Dict[str, Any]] = [{"classId": {"$in": class_ids}}]
    if class_oids:
        variants.append({"classId": {"$in": class_oids}})
    return {"$or": variants}


def _fetch_day_records(db, day_start: datetime, day_end: datetime, class_ids: List[str]):
    records = []
    names = set(db.list_collection_names())
    class_filter = _build_class_filter(class_ids)
    for coll_name in ("attendance", "attendances"):
        if coll_name not in names:
            continue
        coll = db.get_collection(coll_name)
        date_filter = {
            "$or": [
                {"date": {"$gte": day_start, "$lt": day_end}},
                {"timestamp": {"$gte": day_start, "$lt": day_end}},
                {"createdAt": {"$gte": day_start, "$lt": day_end}},
                {"updatedAt": {"$gte": day_start, "$lt": day_end}},
            ]
        }
        query = {"$and": [date_filter, class_filter]} if class_filter else date_filter
        records.extend(
            list(
                coll.find(query)
            )
        )
    return records


def _fetch_month_records(db, month_start: datetime, month_end: datetime, class_ids: List[str]):
    records = []
    names = set(db.list_collection_names())
    class_filter = _build_class_filter(class_ids)
    for coll_name in ("attendance", "attendances"):
        if coll_name not in names:
            continue
        coll = db.get_collection(coll_name)
        date_filter = {
            "$or": [
                {"date": {"$gte": month_start, "$lt": month_end}},
                {"timestamp": {"$gte": month_start, "$lt": month_end}},
                {"createdAt": {"$gte": month_start, "$lt": month_end}},
                {"updatedAt": {"$gte": month_start, "$lt": month_end}},
            ]
        }
        query = {"$and": [date_filter, class_filter]} if class_filter else date_filter
        records.extend(
            list(
                coll.find(query)
            )
        )
    return records


def export_daily_attendance(db, class_name: str, date: str) -> Path:
    day_start = datetime.strptime(date, "%Y-%m-%d")
    day_end = day_start + timedelta(days=1)

    class_ids = _resolve_class_ids(db, class_name)
    students = _fetch_students_for_class(db, class_name, class_ids)
    student_ids = {_as_str_id(s.get("_id") or s.get("id")) for s in students}

    raw_records = _fetch_day_records(db, day_start, day_end, class_ids)

    # If roster lookup failed but attendance exists, derive roster from attendance records.
    if not students and raw_records:
        derived_students: Dict[str, Dict[str, Any]] = {}
        for rec in raw_records:
            sid = _as_str_id(rec.get("student_id") or rec.get("studentId"))
            if not sid:
                continue
            if sid in derived_students:
                continue
            fallback_name = rec.get("name") or sid
            derived_students[sid] = {
                "_id": sid,
                "name": fallback_name,
                "rollNo": "-",
            }
        students = list(derived_students.values())
        student_ids = set(derived_students.keys())

    by_student: Dict[str, Dict[str, Any]] = {}
    for rec in raw_records:
        sid = _as_str_id(rec.get("student_id") or rec.get("studentId"))
        if sid not in student_ids:
            continue

        status = _normalize_status(rec)
        first_seen = _extract_time(rec.get("first_seen") or rec.get("timestamp"))
        last_seen = _extract_time(rec.get("last_seen") or rec.get("timestamp"))

        current = by_student.get(sid)
        if current is None:
            by_student[sid] = {
                "status": status,
                "first_seen": first_seen,
                "last_seen": last_seen,
            }
        else:
            if status == "Present":
                current["status"] = "Present"
            if current["first_seen"] == "-" and first_seen != "-":
                current["first_seen"] = first_seen
            if last_seen != "-":
                current["last_seen"] = last_seen

    rows: List[Dict[str, Any]] = []
    for stu in students:
        sid = _as_str_id(stu.get("_id") or stu.get("id"))
        info = by_student.get(sid, {})
        rows.append(
            {
                "Student ID": sid or "-",
                "Name": stu.get("name") or "-",
                "Roll No": stu.get("rollNo") or "-",
                "Class": class_name,
                "Date": date,
                "Status": info.get("status", "Absent"),
                "First Seen": info.get("first_seen", "-"),
                "Last Seen": info.get("last_seen", "-"),
            }
        )

    if not rows:
        rows.append({
            "Student ID": "-",
            "Name": "-",
            "Class": class_name,
            "Date": date,
            "Status": "-",
            "First Seen": "-",
            "Last Seen": "-",
        })

    df = pd.DataFrame(rows).fillna("-").sort_values(by="Name", kind="stable")

    total_students = len([r for r in rows if r["Student ID"] != "-"])
    present_count = sum(1 for r in rows if r["Status"] == "Present")
    absent_count = sum(1 for r in rows if r["Status"] == "Absent")
    percentage = round((present_count / total_students) * 100, 2) if total_students else 0.0

    summary_df = pd.DataFrame(
        [
            {"Metric": "Total Students", "Value": total_students},
            {"Metric": "Present Count", "Value": present_count},
            {"Metric": "Absent Count", "Value": absent_count},
            {"Metric": "Attendance Percentage", "Value": f"{percentage}%"},
        ]
    )

    file_path = EXPORT_DIR / f"attendance_{_safe_class_name(class_name)}_{date}.xlsx"
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Daily Report")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    _apply_sheet_style(wb["Daily Report"])
    _apply_sheet_style(wb["Summary"])
    wb.save(file_path)
    return file_path


def export_monthly_attendance(db, class_name: str, month: int, year: int) -> Path:
    month = int(month)
    year = int(year)
    month_start = datetime(year, month, 1)
    month_end = (month_start + timedelta(days=32)).replace(day=1)
    days_in_month = calendar.monthrange(year, month)[1]

    class_ids = _resolve_class_ids(db, class_name)
    students = _fetch_students_for_class(db, class_name, class_ids)
    student_ids = {_as_str_id(s.get("_id") or s.get("id")) for s in students}

    month_records = _fetch_month_records(db, month_start, month_end, class_ids)
    day_cols = [f"Day {d}" for d in range(1, days_in_month + 1)]

    if not students and month_records:
        derived_students: Dict[str, Dict[str, Any]] = {}
        for rec in month_records:
            sid = _as_str_id(rec.get("student_id") or rec.get("studentId"))
            if not sid:
                continue
            if sid in derived_students:
                continue
            derived_students[sid] = {
                "_id": sid,
                "name": rec.get("name") or sid,
                "rollNo": "-",
            }
        students = list(derived_students.values())
        student_ids = set(derived_students.keys())

    day_map: Dict[str, Dict[str, str]] = {}
    for stu in students:
        sid = _as_str_id(stu.get("_id") or stu.get("id"))
        day_map[sid] = {f"Day {d}": "A" for d in range(1, days_in_month + 1)}

    for rec in month_records:
        sid = _as_str_id(rec.get("student_id") or rec.get("studentId"))
        if sid not in student_ids:
            continue
        day_str = _extract_day(rec.get("date") or rec.get("timestamp"))
        if not day_str:
            continue
        try:
            day_num = int(day_str[-2:])
        except Exception:
            continue
        if 1 <= day_num <= days_in_month:
            day_map[sid][f"Day {day_num}"] = "P" if _normalize_status(rec) == "Present" else "A"

    monthly_rows: List[Dict[str, Any]] = []
    summary_rows: List[Dict[str, Any]] = []
    for stu in students:
        sid = _as_str_id(stu.get("_id") or stu.get("id"))
        name = stu.get("name") or sid
        row = {"Name": name, "Roll No": stu.get("rollNo") or "-"}
        row.update(day_map.get(sid, {c: "A" for c in day_cols}))

        present_days = sum(1 for c in day_cols if row[c] == "P")
        absent_days = days_in_month - present_days
        pct = round((present_days / float(days_in_month)) * 100, 2) if days_in_month else 0.0

        row["Present Days"] = present_days
        row["Attendance %"] = f"{pct}%"
        monthly_rows.append(row)

        summary_rows.append(
            {
                "Name": name,
                "Present Count": present_days,
                "Absent Count": absent_days,
                "Percentage": f"{pct}%",
            }
        )

    if not monthly_rows:
        row = {"Name": "-"}
        row.update({c: "-" for c in day_cols})
        row["Present Days"] = "-"
        row["Attendance %"] = "-"
        monthly_rows.append(row)
        summary_rows.append({"Name": "-", "Present Count": "-", "Absent Count": "-", "Percentage": "-"})

    monthly_df = pd.DataFrame(monthly_rows).fillna("-").sort_values(by="Name", kind="stable")
    summary_df = pd.DataFrame(summary_rows).fillna("-").sort_values(by="Name", kind="stable")

    file_path = EXPORT_DIR / f"attendance_{_safe_class_name(class_name)}_{year}_{month:02d}.xlsx"
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        monthly_df.to_excel(writer, index=False, sheet_name="Monthly Report")
        summary_df.to_excel(writer, index=False, sheet_name="Student Summary")

    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    _apply_sheet_style(wb["Monthly Report"])
    _apply_sheet_style(wb["Student Summary"])
    wb.save(file_path)
    return file_path
